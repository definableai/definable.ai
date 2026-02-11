"""XLSX parser — extracts text from Excel spreadsheets using openpyxl.

Requires optional dependency: ``openpyxl>=3.1.0``
"""

import io
from typing import List, Set

from definable.readers.models import ContentBlock, ReaderConfig
from definable.readers.parsers.base_parser import BaseParser


def _import_openpyxl():
  try:
    import openpyxl

    return openpyxl
  except ImportError as e:
    raise ImportError(
      "XlsxParser requires the 'openpyxl' package. Install it with: pip install 'openpyxl>=3.1.0' or: pip install 'definable[readers]'"
    ) from e


class XlsxParser(BaseParser):
  """Parses XLSX files using openpyxl (local, no API calls).

  Extracts sheets as tab-separated text with a guard against
  very large spreadsheets.
  """

  def __init__(self, max_rows: int = 1000) -> None:
    _import_openpyxl()
    self.max_rows = max_rows

  def supported_mime_types(self) -> List[str]:
    return ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]

  def supported_extensions(self) -> Set[str]:
    return {".xlsx"}

  def parse(self, data: bytes, *, mime_type: str | None = None, config: ReaderConfig | None = None) -> List[ContentBlock]:
    openpyxl = _import_openpyxl()
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    blocks: List[ContentBlock] = []

    for sheet_name in wb.sheetnames:
      ws = wb[sheet_name]
      rows: List[str] = []
      row_count = 0
      for row in ws.iter_rows(values_only=True):
        if row_count >= self.max_rows:
          break
        cells = [str(cell) if cell is not None else "" for cell in row]
        rows.append("\t".join(cells))
        row_count += 1

      if rows:
        sheet_text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
        blocks.append(
          ContentBlock(
            content_type="table",
            content=sheet_text,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata={"sheet_name": sheet_name},
          )
        )

    wb.close()
    return blocks
